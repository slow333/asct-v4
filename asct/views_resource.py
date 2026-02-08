from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from .models_resource import CPUUsage, MemoryUsage, NetworkUsage
from .models_basic import SSHInfo
from .run_by_ssh import run_ssh_cpu_usage, run_ssh_memory_usage, run_ssh_traffic_usage
import openpyxl
from django.utils import timezone
from datetime import timedelta
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
import base64

# =============== Traffic usage 관련 CRUD ===============
@login_required
def traffic_usage_select(request):
    if request.method == 'POST':
        ssh_id = request.POST.get('ssh_id')
        if ssh_id:
            return redirect('asct:traffic_usage_run', ssh_id=ssh_id)
    sshinfos = SSHInfo.objects.filter(operators=request.user)
    return render(request, 'asct/traffic_usage/select.html', {'sshinfos': sshinfos})

@login_required
def traffic_usage_run(request, ssh_id):
    ssh_info = get_object_or_404(SSHInfo, id=ssh_id)
    _, _, data, error = run_ssh_traffic_usage(request, ssh_info)
    
    if error:
        messages.error(request, f"Error: {error}")
    else:
        messages.success(request, f"Successfully collected {data.get('count', 0)} records.")
    return redirect('asct:traffic_usage_list')

def traffic_usage_list(request):
    query = request.GET.get('q', '')
    host_list = NetworkUsage.objects.exclude(hostname__isnull=True).values_list('hostname', flat=True).distinct().order_by('hostname')

    if query:
        network_usage = NetworkUsage.objects.filter(hostname=query)
    else:
        network_usage = NetworkUsage.objects.all()
    
    pagenator = Paginator(network_usage, 10)
    page = request.GET.get("page")
    page_obj = pagenator.get_page(page)
    
    return render(request, 'asct/traffic_usage/list.html', {'page_obj': page_obj, 'query': query, 'host_list': host_list})

@login_required
def traffic_usage_export(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="traffic_usage_list.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traffic Usage" # type: ignore

    headers = ['Hostname', 'IP', 'Date Time', 'Interface', 'Speed', 'RX(kB/s)', 'TX(kB/s)', 'Confirmed', 'Comment']
    ws.append(headers) # type: ignore

    traffics = NetworkUsage.objects.all().order_by('hostname', '-data_time')
    for t in traffics:
        data_time_val = t.data_time.replace(tzinfo=None) if t.data_time else ''
        ws.append([ # type: ignore
            t.hostname,
            t.ip,
            data_time_val,
            t.if_name,
            t.speed,
            t.rxkB_s,
            t.txkB_s,
            "Yes" if t.is_confirmed else "No",
            t.comment
        ])

    wb.save(response)
    return response

@login_required
def traffic_usage_chart(request):
    period = request.GET.get('period', '1m')
    query = request.GET.get('q', '')
    
    host_list = NetworkUsage.objects.exclude(hostname__isnull=True).values_list('hostname', flat=True).distinct().order_by('hostname')
    
    queryset = NetworkUsage.objects.all().order_by('data_time')
    
    if query:
        queryset = queryset.filter(hostname=query)
    
    if period == '1w':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=7))
    elif period == '1m':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=30))
    elif period == '3m':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=90))
        
    fig, ax = plt.subplots(figsize=(12, 6))
    
    data_map = {}
    for entry in queryset:
        # 인터페이스별 RX/TX 구분
        rx_key = f"{entry.hostname} - {entry.if_name} (RX)"
        tx_key = f"{entry.hostname} - {entry.if_name} (TX)"
        
        if rx_key not in data_map: data_map[rx_key] = {'x': [], 'y': []}
        if tx_key not in data_map: data_map[tx_key] = {'x': [], 'y': []}
            
        data_map[rx_key]['x'].append(entry.data_time)
        data_map[rx_key]['y'].append(float(entry.rxkB_s))
        
        data_map[tx_key]['x'].append(entry.data_time)
        data_map[tx_key]['y'].append(float(entry.txkB_s))

    for label, data in data_map.items():
        ax.plot(data['x'], data['y'], label=label, marker='o', markersize=3)
        
    ax.set_title(f'Traffic Usage ({period})')
    ax.set_xlabel('Date Time')
    ax.set_ylabel('Speed (kB/s)')
    
    # 범례가 너무 많으면 가독성을 해치므로 데이터가 적을 때만 표시
    if len(data_map) > 0 and len(data_map) < 20:
        ax.legend()
        
    ax.grid(True)
    fig.tight_layout()
    
    buffer = io.BytesIO()
    fig.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()
    plt.close(fig)
    
    graphic = base64.b64encode(image_png).decode('utf-8')
    
    context = {
        'chart_graphic': graphic,
        'period': period,
        'query': query,
        'host_list': host_list
    }
    return render(request, 'asct/traffic_usage/chart.html', context)

# =============== Memory usage 관련 CRUD ===============
@login_required
def memory_usage_list(request):
    query = request.GET.get('q', '')
    host_list = MemoryUsage.objects.exclude(hostname__isnull=True).values_list('hostname', flat=True).distinct().order_by('hostname')

    if query:
        memory_usage = MemoryUsage.objects.filter(hostname=query)
    else:
        memory_usage = MemoryUsage.objects.all()
    
    pagenator = Paginator(memory_usage, 10)
    page = request.GET.get("page")
    page_obj = pagenator.get_page(page)
    
    return render(request, 'asct/memory_usage/list.html', {'page_obj': page_obj, 'query': query, 'host_list': host_list})

@login_required
def memory_usage_chart(request):
    period = request.GET.get('period', '1m')
    query = request.GET.get('q', '')
    host_list = MemoryUsage.objects.exclude(hostname__isnull=True).values_list('hostname', flat=True).distinct().order_by('hostname')
    
    queryset = MemoryUsage.objects.all().order_by('data_time')
    
    if query:
        queryset = queryset.filter(hostname=query)
    
    if period == '1w':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=7))
    elif period == '1m':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=30))
    elif period == '3m':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=90))
    
    # Matplotlib 설정
    fig, ax = plt.subplots(figsize=(12, 6))
    
    data_map = {}
    for entry in queryset:
        host = entry.hostname
        if host not in data_map:
            data_map[host] = {'x': [], 'y': []}
        data_map[host]['x'].append(entry.data_time)
        data_map[host]['y'].append(float(entry.usage_p))
        
    for host, data in data_map.items():
        ax.plot(data['x'], data['y'], label=host, marker='o', markersize=3)

    ax.set_title(f'Memory Usage ({period})')
    ax.set_xlabel('Date Time')
    ax.set_ylabel('Usage (%)')
    if data_map:
        ax.legend()
    ax.grid(True)
    fig.tight_layout()
    
    buffer = io.BytesIO()
    fig.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()
    plt.close(fig)
    
    graphic = base64.b64encode(image_png).decode('utf-8')
    
    context = {
        'chart_graphic': graphic,
        'period': period,
        'query': query,
        'host_list': host_list
    }
    return render(request, 'asct/memory_usage/chart.html', context)

@login_required
def memory_usage_select(request):
    if request.method == 'POST':
        ssh_id = request.POST.get('ssh_id')
        if ssh_id:
            return redirect('asct:memory_usage_run', ssh_id=ssh_id)
    sshinfos = SSHInfo.objects.filter(operators=request.user)
    return render(request, 'asct/memory_usage/select.html', {'sshinfos': sshinfos})

@login_required
def memory_usage_run(request, ssh_id):
    ssh_info = get_object_or_404(SSHInfo, id=ssh_id)
    _, _, data, error = run_ssh_memory_usage(request, ssh_info)
    
    if error:
        messages.error(request, f"Error: {error}")
    else:
        messages.success(request, f"Successfully collected {data.get('count', 0)} records.")
    return redirect('asct:memory_usage_list')

@login_required
def memory_usage_export(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="memory_usage_list.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Memory Usage" # type: ignore
    headers = ['Hostname', 'IP', 'Date Time', 'Total Memory(MB)', 'Usage(%)', 'Confirmed', 'Comment']
    ws.append(headers) # type: ignore
    mem_usages = MemoryUsage.objects.all().order_by('hostname', '-data_time')
    for mem in mem_usages:
        data_time_val = mem.data_time.replace(tzinfo=None) if mem.data_time else ''
        ws.append([ # type: ignore
            mem.hostname, mem.ip, data_time_val, mem.total_memory, mem.usage_p,
            "Yes" if mem.is_confirmed else "No", mem.comment
        ])
    wb.save(response)
    return response

# =============== CPU usage 관련 CRUD ===============
@login_required
def cpu_usage_list(request):
    query = request.GET.get('q', '')
    
    # Dropdown을 위한 호스트 목록 조회 (중복 제거)
    host_list = CPUUsage.objects.exclude(hostname__isnull=True).values_list('hostname', flat=True).distinct().order_by('hostname')

    if query:
        cpu_usage = CPUUsage.objects.filter(hostname=query)
    else:
        cpu_usage = CPUUsage.objects.all()
    
    pagenator = Paginator(cpu_usage, 10)
    page = request.GET.get("page")
    page_obj = pagenator.get_page(page)
    
    return render(request, 'asct/cpu_usage/list.html', {'page_obj': page_obj, 'query': query, 'host_list': host_list})

@login_required
def cpu_usage_chart(request):
    period = request.GET.get('period', '1m')
    query = request.GET.get('q', '')
    
    # Dropdown을 위한 호스트 목록 조회 (중복 제거)
    host_list = CPUUsage.objects.exclude(hostname__isnull=True).values_list('hostname', flat=True).distinct().order_by('hostname')
    
    # 날짜순으로 전체 데이터 조회
    queryset = CPUUsage.objects.all().order_by('data_time')
    
    if query:
        queryset = queryset.filter(hostname=query)
    
    if period == '1w':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=7))
    elif period == '1m':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=30))
    elif period == '3m':
        queryset = queryset.filter(data_time__gte=timezone.now() - timedelta(days=90))
    
    # Matplotlib 설정
    fig, ax = plt.subplots(figsize=(12, 6))
    
    data_map = {}
    for entry in queryset:
        host = entry.hostname
        if host not in data_map:
            data_map[host] = {'x': [], 'y': []}
        data_map[host]['x'].append(entry.data_time)
        data_map[host]['y'].append(float(entry.usage_p))

    for host, data in data_map.items():
        ax.plot(data['x'], data['y'], label=host, marker='o', markersize=3)
    
    ax.set_title(f'CPU Usage ({period})')
    ax.set_xlabel('Date Time')
    ax.set_ylabel('Usage (%)')
    if data_map:
        ax.legend()
    ax.grid(True)
    fig.tight_layout()
    
    buffer = io.BytesIO()
    fig.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()
    plt.close(fig)
    
    graphic = base64.b64encode(image_png).decode('utf-8')

    context = {
        'chart_graphic': graphic,
        'period': period,
        'query': query,
        'host_list': host_list
    }
    return render(request, 'asct/cpu_usage/chart.html', context)

# =============== Paramiko 실행 예시 ===============
@login_required
def cpu_usage_select(request):
    if request.method == 'POST':
        ssh_id = request.POST.get('ssh_id')
        if ssh_id:
            return redirect('asct:cpu_usage_run', ssh_id=ssh_id)

    # 현재 로그인한 사용자가 권한을 가진 서버만 조회
    sshinfos = SSHInfo.objects.filter(operators=request.user)
    
    return render(request, 'asct/cpu_usage/select.html', {'sshinfos': sshinfos})

@login_required
def cpu_usage_run(request, ssh_id):
    ssh_info = get_object_or_404(SSHInfo, id=ssh_id)
    _, _, data, error = run_ssh_cpu_usage(request, ssh_info)
    
    if error:
        messages.error(request, f"Error: {error}")
    else:
        messages.success(request, f"Successfully collected {data.get('count', 0)} records.")
        
    return redirect('asct:cpu_usage_list')

@login_required
def cpu_usage_export(request):
    # 1. 응답 객체 생성 (Excel 파일 설정)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cpu_usage_list.xlsx"'

    # 2. 워크북 및 워크시트 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPU Usage" # type: ignore

    # 3. 헤더 작성
    headers = ['Hostname', 'IP', 'Date Time', 'CPU Cores', 'Usage(%)', 'Confirmed', 'Comment']
    ws.append(headers) # type: ignore

    # 4. 데이터 작성
    cpu_usages = CPUUsage.objects.all().order_by('hostname', '-data_time')
    for cpu in cpu_usages:
        # Timezone 정보가 있는 datetime은 Excel 호환성을 위해 tzinfo 제거
        data_time_val = cpu.data_time.replace(tzinfo=None) if cpu.data_time else ''
        
        ws.append([ # type: ignore
            cpu.hostname,
            cpu.ip,
            data_time_val,
            cpu.cpu_cores,
            cpu.usage_p,
            "Yes" if cpu.is_confirmed else "No",
            cpu.comment
        ])

    # 5. 저장 및 반환
    wb.save(response)
    return response